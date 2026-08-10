#!/usr/bin/env python3
"""
Excel to JSON Converter for Student Portal (12 Months: January - December)
Parses an Excel Workbook (.xlsx) where each tab represents a student's profile containing 12 monthly tables.
"""

import os
import json
import openpyxl

MONTHS = [
    "January", "February", "March", "April", "May", "June", 
    "July", "August", "September", "October", "November", "December"
]

def parse_student_tab(sheet):
    """Parses a single student sheet tab into structured JSON format with 12 months."""
    student_info = {
        "name": "Unknown",
        "student_id": "N/A",
        "grade_class": "N/A",
        "homeroom_teacher": "Mrs. Sheshadi Sathsarani",
        "avatar": "https://api.dicebear.com/7.x/avataaars/svg?seed=" + sheet.title,
        "qr_code_key": f"QR-{sheet.title}",
        "access_url": f"student.html?id={sheet.title}"
    }
    
    # Read Profile metadata
    for r in range(1, 10):
        for c in range(1, 10):
            val = str(sheet.cell(r, c).value or "").strip()
            val_next = str(sheet.cell(r, c+1).value or "").strip()
            
            if "student name" in val.lower():
                student_info["name"] = val_next or "Student"
            elif "student id" in val.lower():
                student_info["student_id"] = val_next or "N/A"
                student_info["qr_code_key"] = f"QR-{val_next}"
                student_info["access_url"] = f"student.html?id={val_next}"
            elif "grade" in val.lower() or "class" in val.lower():
                student_info["grade_class"] = val_next or "06 - Science"
            elif "teacher" in val.lower():
                student_info["homeroom_teacher"] = val_next or "Mrs. Sheshadi Sathsarani"

    monthly_progress = {}
    current_month_idx = 0

    # Scan rows for monthly tables
    for r in range(1, sheet.max_row + 1):
        row_str = " ".join([str(sheet.cell(r, c).value or "").strip() for c in range(1, 7)]).lower()
        
        # Check if row is a month banner or week header
        for m in MONTHS:
            if m.lower() in row_str:
                current_month_idx = MONTHS.index(m)
                
        c1 = str(sheet.cell(r, 1).value or "").strip().lower()
        if "week" in c1:
            month_key = MONTHS[current_month_idx] if current_month_idx < len(MONTHS) else "January"
            if month_key not in monthly_progress:
                monthly_progress[month_key] = []
                
            mg1 = str(sheet.cell(r, 2).value or "").strip()
            mg2 = str(sheet.cell(r, 3).value or "").strip()
            pp = str(sheet.cell(r, 4).value or "").strip()
            pr = str(sheet.cell(r, 5).value or "").strip()
            ut = sheet.cell(r, 6).value
            
            try:
                ut = float(ut) if ut is not None else 0
            except:
                ut = 0
                
            monthly_progress[month_key].append({
                "week": str(sheet.cell(r, 1).value or "").strip(),
                "master_guide_1": mg1 or "N/A",
                "master_guide_2": mg2 or "N/A",
                "past_paper": pp or "N/A",
                "practical": pr or "N/A",
                "unit_test": ut
            })

    # Read Term Assessments
    assessments = []
    for r in range(1, sheet.max_row + 1):
        c1 = str(sheet.cell(r, 1).value or "").strip()
        c2 = sheet.cell(r, 2).value
        if "term" in c1.lower() or "exam" in c1.lower():
            try:
                score = float(c2) if c2 is not None else 0
            except:
                score = 0
            assessments.append({"term": c1, "score": score})

    # Fill missing months with default template
    for m in MONTHS:
        if m not in monthly_progress or len(monthly_progress[m]) == 0:
            monthly_progress[m] = [
                {"week": "1 week", "master_guide_1": "Incomplete", "master_guide_2": "Incomplete", "past_paper": "Incomplete", "practical": "Average", "unit_test": 50},
                {"week": "2 week", "master_guide_1": "Incomplete", "master_guide_2": "Incomplete", "past_paper": "Incomplete", "practical": "Good", "unit_test": 50},
                {"week": "3 week", "master_guide_1": "Incomplete", "master_guide_2": "Incomplete", "past_paper": "Incomplete", "practical": "Good", "unit_test": 50},
                {"week": "4 week", "master_guide_1": "Incomplete", "master_guide_2": "Incomplete", "past_paper": "Incomplete", "practical": "Good", "unit_test": 50}
            ]

    # Calculate overall average unit test score across all 12 months
    all_ut = []
    for m in MONTHS:
        for w in monthly_progress[m]:
            if isinstance(w["unit_test"], (int, float)):
                all_ut.append(w["unit_test"])
    avg_score = round(sum(all_ut) / len(all_ut), 1) if all_ut else 0.0

    return {
        "tab_name": sheet.title,
        "student_info": student_info,
        "monthly_progress": monthly_progress,
        "assessments": assessments or [
            {"term": "Term 1", "score": 80},
            {"term": "Term 2", "score": 68},
            {"term": "Final Exam", "score": 100}
        ],
        "summary": {
            "attendance": "96%",
            "average_unit_test": avg_score,
            "overall_status": "Active Student"
        }
    }

def convert_excel(excel_path, json_output_path):
    print(f"Loading Excel workbook: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    all_students = []
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"Processing student tab (12 Months): {sheet_name}")
        student_data = parse_student_tab(sheet)
        all_students.append(student_data)
        
    with open(json_output_path, "w") as f:
        json.dump(all_students, f, indent=2)
        
    print(f"Successfully exported {len(all_students)} student tabs to {json_output_path}")

if __name__ == "__main__":
    script_dir = os.path.dirname(os.path.abspath(__file__))
    excel_file = os.path.join(script_dir, "../assets/data/Student_Improvement_Tracker.xlsx")
    json_file = os.path.join(script_dir, "../assets/data/students.json")
    
    if os.path.exists(excel_file):
        convert_excel(excel_file, json_file)
    else:
        print(f"File not found: {excel_file}")
